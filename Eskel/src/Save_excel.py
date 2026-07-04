try:
    from config.config_setup import setup_logging
    data_file = setup_logging()
except ModuleNotFoundError as e:
    print(f'Error: Module not found {e}')
    raise SystemExit(1)
import logging
logger = logging.getLogger(__name__)
try:
    from openpyxl import Workbook
except ModuleNotFoundError as e:
        logger.error(f'Error: Module not found {e}')
        raise SystemExit(1)
try:
    import pandas as pd
except ModuleNotFoundError as e:
        logger.error(f'Error: Module not found {e}')
        raise SystemExit(1)
try:
    from reports import Raports
except ModuleNotFoundError as e:
        logger.error(f'Error: Module not found {e}')
        raise SystemExit(1)
try:
    def SaveExcel():
        logger.info('Saving to Excel format...')
        reports = Raports()

        r1 = reports.raport1()
        wb = Workbook()
        ws = wb.active
        ws.title = 'report1'
        for i in r1:
            for key, value in i.items():
                ws.append([key, value])
        report1 = data_file / 'report1.xlsx'
        report1.parent.mkdir(parents=True, exist_ok=True)
        wb.save(report1)

        r2 = reports.raport2()
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "report2"
        ws2['A1'], ws2['B1'] = 'quantity', 'title'
        for index, row in r2.iterrows():
            for i, value in enumerate(row, start=1):
                ws2.cell(row=index+2, column=i, value=value)
        report2 = data_file / 'report2.xlsx'
        report2.parent.mkdir(parents=True, exist_ok=True)
        wb2.save(report2)

        report3 = data_file / 'report3.xlsx'
        report3.parent.mkdir(parents=True, exist_ok=True)
        pd.DataFrame(reports.raport3()).to_excel(report3, index=False, sheet_name='report3')

        report4 = data_file / 'report4.xlsx'
        report4.parent.mkdir(parents=True, exist_ok=True)
        pd.DataFrame(reports.raport4()).to_excel(report4, index=False, sheet_name='report4')

        report5 = data_file / 'report5.xlsx'
        report5.parent.mkdir(parents=True, exist_ok=True)
        reports.raport5().to_excel(report5, index=False, sheet_name='report5')

        report6 = data_file / 'report6.xlsx'
        report6.parent.mkdir(parents=True, exist_ok=True)
        reports.raport6().to_excel(report6, index=False, sheet_name='report6')

        report7 = data_file / 'report7.xlsx'
        report7.parent.mkdir(parents=True, exist_ok=True)
        reports.raport7().to_excel(report7, index=False, sheet_name='report7')

        logger.info('Preparing combined raport')
        AllReports = data_file / 'AllRaports.xlsx'
        AllReports.parent.mkdir(parents=True, exist_ok=True)
        wb.save(AllReports)
        with pd.ExcelWriter(AllReports, mode='a', engine='openpyxl', if_sheet_exists='replace') as f:
            reports.raport2().to_excel(f, index=False, sheet_name='report2')
            pd.DataFrame(reports.raport3()).to_excel(f, index=False, sheet_name='report3')
            pd.DataFrame(reports.raport4()).to_excel(f, index=False, sheet_name='report4')
            reports.raport5().to_excel(f, index=False, sheet_name='report5')
            reports.raport6().to_excel(f, index=False, sheet_name='report6')
            reports.raport7().to_excel(f, index=False, sheet_name='report7')
        logger.info('All excel reports completed')
except PermissionError as e:
    logger.error(f'Error: Lack of permision to save in this file "{e}"')
except FileNotFoundError as e:
    logger.error(f'Error: Path does not exists "{e}"')
except Exception as e:
    logger.error(f'Unexpected error: {e}')

if __name__ == '__main__':
    SaveExcel()